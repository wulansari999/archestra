"""Verify the agent's sales-bonus workbook: real data from the seeded skill, live formulas, correct
per-team totals.

The source ledger lives only in the mounted `sales-ledger` skill, so the data-fidelity check against
the never-staged ground truth (BENCH_FIXTURES/expected/sales.json) plus a BENCH_STATE check that the
skill was loaded and read is the skill-gate. The bonus and per-team formulas must be live formulas
referencing the right cells (openpyxl cannot evaluate them, so per-team values are cross-checked
against a recompute from sales.json); a workbook with hardcoded numbers and no/degenerate formulas
fails.
"""

import json
import os
import re
from pathlib import Path

import openpyxl

HEADER = ["sale_id", "salesperson", "team", "units", "unit_price", "commission_pct", "bonus"]
FIRST_DATA_ROW = 2
SKILL_MOUNT = "/skills/sales-ledger"
FLOAT_TOL = 1e-9
MONEY_TOL = 0.01

CELL_RE = re.compile(r"^[A-Z]+[0-9]+$")
RANGE_RE = re.compile(r"^[A-Z]+[0-9]+:[A-Z]+[0-9]+$")


# === fixtures / state loading ===

def _load_json(env_var: str) -> dict:
    path = os.environ.get(env_var)
    assert path, f"{env_var} is not set"
    return json.loads(Path(path).read_text(encoding="utf-8"))


def _sales() -> list[dict]:
    base = os.environ.get("BENCH_FIXTURES")
    assert base, "BENCH_FIXTURES is not set"
    return json.loads(Path(base, "expected", "sales.json").read_text(encoding="utf-8"))


def _workbook():
    path = os.environ.get("BENCH_OUTPUT")
    assert path, "BENCH_OUTPUT is not set -- the agent did not export a workbook"
    ws = openpyxl.load_workbook(path, data_only=False).active
    return ws


# === formula normalization ===

def _strip_qualifier(token: str, sheet_title: str) -> str:
    token = token.strip()
    if "!" not in token:
        return token
    qual, ref = token.rsplit("!", 1)
    qual = qual.strip()
    if qual.startswith("'") and qual.endswith("'"):
        qual = qual[1:-1].replace("''", "'")
    assert qual == sheet_title, f"cross-sheet reference not allowed: {token!r}"
    return ref


def _norm_ref(token: str, sheet_title: str) -> str:
    ref = _strip_qualifier(token, sheet_title).replace("$", "").upper()
    assert CELL_RE.match(ref), f"not a plain cell reference: {token!r}"
    return ref


def _norm_range(token: str, sheet_title: str) -> str:
    rng = _strip_qualifier(token, sheet_title).replace("$", "").upper()
    assert RANGE_RE.match(rng), f"not a plain range: {token!r}"
    return rng


def _split_top_level(text: str) -> list[str]:
    args, depth, cur = [], 0, ""
    for ch in text:
        if ch == "(":
            depth += 1
            cur += ch
        elif ch == ")":
            depth -= 1
            cur += ch
        elif ch == "," and depth == 0:
            args.append(cur)
            cur = ""
        else:
            cur += ch
    args.append(cur)
    return args


def _product_operands(formula, sheet_title: str) -> set[str]:
    assert isinstance(formula, str) and formula.startswith("="), (
        f"expected a live formula, got {formula!r} (a hardcoded value fails)"
    )
    body = formula[1:]
    assert "(" not in body and ")" not in body, f"unexpected call in bonus formula: {formula!r}"
    return {_norm_ref(p, sheet_title) for p in body.split("*")}


def _parse_sumif(formula, sheet_title: str) -> tuple[str, str, str]:
    assert isinstance(formula, str) and formula.startswith("="), (
        f"expected a SUMIF formula, got {formula!r} (a hardcoded value fails)"
    )
    f = formula.strip()
    assert f.upper().startswith("=SUMIF(") and f.endswith(")"), f"team total must be SUMIF: {formula!r}"
    args = _split_top_level(f[len("=SUMIF(") : -1])
    assert len(args) == 3, f"SUMIF needs 3 args: {formula!r}"
    return _norm_range(args[0], sheet_title), args[1].strip(), _norm_range(args[2], sheet_title)


def _resolve_criterion(crit: str, ws, sheet_title: str):
    crit = crit.strip()
    if len(crit) >= 2 and crit[0] == '"' and crit[-1] == '"':
        return crit[1:-1]
    return ws[_norm_ref(crit, sheet_title)].value


# === checks ===

def test_skill_seeded_from_repo() -> None:
    rest = _load_json("BENCH_STATE")["rest"]
    assert len(rest) == 1, f"expected one captured rest path, got {list(rest)}"
    payload = next(iter(rest.values()))
    rows = payload.get("data") if isinstance(payload, dict) else None
    matches = [r for r in (rows or []) if r.get("name") == "sales-ledger"]
    assert len(matches) == 1, f"expected exactly one sales-ledger skill row, got {matches}"
    assert matches[0].get("sourceType") == "github", (
        f"sales-ledger must be seeded from the repo (sourceType=github), got {matches[0].get('sourceType')!r}"
    )


def test_skill_loaded_and_read() -> None:
    calls = _load_json("BENCH_STATE").get("tool_calls", [])
    loaded = [
        c for c in calls
        if c.get("name") == "archestra__load_skill"
        and "sales-ledger" in json.dumps(c.get("input") or {})
    ]
    assert loaded, "no load_skill call for sales-ledger; the skill was not loaded"
    read = [
        c for c in calls
        if c.get("name") == "archestra__run_command"
        and SKILL_MOUNT in " ".join(str((c.get("input") or {}).get(k) or "") for k in ("command", "cwd"))
    ]
    assert read, f"no run_command referenced the mounted {SKILL_MOUNT}; the ledger was not read"


def test_data_fidelity() -> None:
    ws, sales = _workbook(), _sales()
    header = [ws.cell(row=1, column=c).value for c in range(1, 7)]
    assert header == HEADER[:6], f"header A1:F1 must be {HEADER[:6]}, got {header}"
    for idx, row in enumerate(sales):
        r = FIRST_DATA_ROW + idx
        got = {
            "sale_id": ws.cell(row=r, column=1).value,
            "salesperson": ws.cell(row=r, column=2).value,
            "team": ws.cell(row=r, column=3).value,
            "units": ws.cell(row=r, column=4).value,
            "unit_price": ws.cell(row=r, column=5).value,
            "commission_pct": ws.cell(row=r, column=6).value,
        }
        for key in ("sale_id", "salesperson", "team", "units"):
            assert got[key] == row[key], f"row {r} {key}: {got[key]!r} != {row[key]!r}"
        for key in ("unit_price", "commission_pct"):
            assert isinstance(got[key], (int, float)) and abs(got[key] - row[key]) < FLOAT_TOL, (
                f"row {r} {key}: {got[key]!r} != {row[key]!r}"
            )
    extra = ws.cell(row=FIRST_DATA_ROW + len(sales), column=1).value
    assert extra is None, f"unexpected extra data row at {FIRST_DATA_ROW + len(sales)}: {extra!r}"


def test_bonus_formulas() -> None:
    ws, sales = _workbook(), _sales()
    title = ws.title
    for idx in range(len(sales)):
        r = FIRST_DATA_ROW + idx
        operands = _product_operands(ws.cell(row=r, column=7).value, title)
        assert operands == {f"D{r}", f"E{r}", f"F{r}"}, (
            f"bonus G{r} must be =D{r}*E{r}*F{r} (units*unit_price*commission_pct), got operands {operands}"
        )


def test_team_summary_formulas() -> None:
    ws, sales = _workbook(), _sales()
    title = ws.title
    last = FIRST_DATA_ROW + len(sales) - 1
    expected_teams = {row["team"] for row in sales}

    seen: list[str] = []
    row = 2
    while True:
        label = ws.cell(row=row, column=9).value  # col I
        if label is None:
            break
        formula = ws.cell(row=row, column=10).value  # col J
        rng1, crit, rng2 = _parse_sumif(formula, title)
        assert rng1 == f"C{FIRST_DATA_ROW}:C{last}", f"SUMIF range1 must be C{FIRST_DATA_ROW}:C{last}, got {rng1}"
        assert rng2 == f"G{FIRST_DATA_ROW}:G{last}", f"SUMIF range2 must be G{FIRST_DATA_ROW}:G{last}, got {rng2}"
        resolved = _resolve_criterion(crit, ws, title)
        assert resolved == label, f"row {row} SUMIF criterion {resolved!r} must match its team label {label!r}"
        seen.append(label)
        row += 1

    assert seen, "no team summary rows found in column I"
    assert len(seen) == len(set(seen)), f"duplicate team in summary: {seen}"
    assert set(seen) == expected_teams, f"summary teams {set(seen)} != expected {expected_teams}"


def test_submitted_team_bonuses() -> None:
    sales = _sales()
    expected: dict[str, float] = {}
    for row in sales:
        bonus = row["units"] * row["unit_price"] * row["commission_pct"]
        expected[row["team"]] = expected.get(row["team"], 0.0) + bonus

    submitted = _load_json("BENCH_RESULT")["team_bonuses"]
    teams = [item["team"] for item in submitted]
    assert len(teams) == len(set(teams)), f"duplicate team in submission: {teams}"
    assert set(teams) == set(expected), f"submitted teams {set(teams)} != expected {set(expected)}"
    for item in submitted:
        got, want = item["total_bonus"], expected[item["team"]]
        assert abs(got - want) <= MONEY_TOL, f"team {item['team']} bonus {got} != expected {want:.4f}"
