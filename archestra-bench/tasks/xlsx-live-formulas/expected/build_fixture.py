#!/usr/bin/env python3
"""Regenerate the xlsx-live-formulas fixtures from one canonical dataset:

- `skills/sales-ledger/assets/ledger.xlsx` -- the source ledger bundled in the seeded `sales-ledger`
  skill and mounted into the sandbox (cols A..F, no bonus). This is the ONLY place the agent can read
  the data; it is never in the prompt.
- `tasks/xlsx-live-formulas/expected/sales.json` -- the verifier's never-staged ground truth.

Both are written from CANONICAL below, so they cannot drift; the script reads ledger.xlsx back and
asserts it matches sales.json. Deterministic: no RNG, no clock (workbook timestamps are pinned).

Run: `uv run --with openpyxl==3.1.5 archestra-bench/tasks/xlsx-live-formulas/expected/build_fixture.py`
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl

# 8 salespeople, 2 per team across 4 teams; each a distinct commission rate.
PEOPLE = [
    ("Ana", "North", 0.035),
    ("Ben", "North", 0.041),
    ("Cara", "East", 0.028),
    ("Drew", "East", 0.052),
    ("Esme", "South", 0.047),
    ("Finn", "South", 0.033),
    ("Gita", "West", 0.039),
    ("Hugo", "West", 0.044),
]
N_SALES = 30
HEADER = ["sale_id", "salesperson", "team", "units", "unit_price", "commission_pct"]
FIXED_TS = datetime(2024, 1, 1, 0, 0, 0)


def canonical() -> list[dict]:
    rows: list[dict] = []
    for i in range(N_SALES):
        name, team, commission = PEOPLE[i % len(PEOPLE)]
        units = 3 + (i * 7) % 40
        unit_price = round(12.5 + (i % 9) * 7.3 + (i % 5) * 1.07, 2)
        rows.append(
            {
                "sale_id": f"S{i + 1:03d}",
                "salesperson": name,
                "team": team,
                "units": units,
                "unit_price": unit_price,
                "commission_pct": commission,
            }
        )
    return rows


def write_ledger(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.properties.created = FIXED_TS
    wb.properties.modified = FIXED_TS
    ws = wb.active
    ws.title = "Ledger"
    ws.append(HEADER)
    for r in rows:
        ws.append([r[k] for k in HEADER])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def read_ledger(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    assert values[0] == tuple(HEADER), f"ledger header mismatch: {values[0]}"
    return [dict(zip(HEADER, row)) for row in values[1:]]


def main() -> None:
    expected_dir = Path(__file__).resolve().parent
    repo_root = expected_dir.parents[3]  # .../archestra-bench/tasks/xlsx-live-formulas/expected
    ledger_path = repo_root / "archestra-bench" / "skills" / "sales-ledger" / "assets" / "ledger.xlsx"
    sales_json_path = expected_dir / "sales.json"

    rows = canonical()
    write_ledger(rows, ledger_path)
    sales_json_path.write_text(json.dumps(rows, indent=2) + "\n", encoding="utf-8")

    readback = read_ledger(ledger_path)
    assert readback == rows, "ledger.xlsx does not match canonical rows (drift)"
    print(f"wrote {len(rows)} sales -> {ledger_path.name} + {sales_json_path.name}; readback OK")


if __name__ == "__main__":
    main()
